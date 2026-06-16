from datetime import datetime
from pathlib import Path

import duckdb
import pandas as pd
import pytest
from openpyxl import Workbook

from churnlens.config.settings import Settings
from churnlens.io.ingest import (
    SOURCE_FILENAME,
    SOURCE_SHEETS,
    ingest_bronze,
    read_source,
)

SOURCE_COLUMNS = [
    "Invoice",
    "StockCode",
    "Description",
    "Quantity",
    "InvoiceDate",
    "Price",
    "Customer ID",
    "Country",
]

# A row present in both sheets — the real UCI sheets overlap, and bronze must keep both.
OVERLAP_ROW = ("491200", "21733", "HEART", 6, datetime(2010, 12, 1, 10, 0), 2.95, 14688, "UK")

SHEET_1_ROWS = [
    ("489434", "85048", "GLASS BALL", 12, datetime(2009, 12, 1, 7, 45), 6.95, 13085, "UK"),
    ("489435", "22350", "CAT BOWL", 8, datetime(2009, 12, 1, 7, 46), 2.55, None, "UK"),
    OVERLAP_ROW,
]
SHEET_2_ROWS = [
    OVERLAP_ROW,
    ("C579889", "23245", "CAKE TINS", -8, datetime(2011, 12, 5, 9, 15), 4.15, 17315, "UK"),
    ("581587", "22138", "BAKING SET", 3, datetime(2011, 12, 9, 12, 50), 4.95, 12680, "France"),
]


def write_workbook(path, sheets):
    wb = Workbook()
    wb.remove(wb.active)
    for name, rows in sheets.items():
        ws = wb.create_sheet(name)
        ws.append(SOURCE_COLUMNS)
        for row in rows:
            ws.append(row)
    wb.save(path)


@pytest.fixture
def settings(tmp_path):
    s = Settings(_env_file=None, data_dir=tmp_path / "data", reports_dir=tmp_path / "reports")
    s.raw_dir.mkdir(parents=True)
    write_workbook(
        s.raw_dir / SOURCE_FILENAME,
        {SOURCE_SHEETS[0]: SHEET_1_ROWS, SOURCE_SHEETS[1]: SHEET_2_ROWS},
    )
    return s


def test_read_source_renames_and_keeps_values(settings):
    frame = read_source(settings.raw_dir / SOURCE_FILENAME)

    assert len(frame) == len(SHEET_1_ROWS) + len(SHEET_2_ROWS)
    assert set(frame.columns) == {
        "invoice",
        "stock_code",
        "description",
        "quantity",
        "invoice_date",
        "unit_price",
        "customer_id",
        "country",
        "source_sheet",
        "source_file",
    }
    # Customer ids are text, never floats; anonymous rows stay null.
    assert frame["customer_id"].iloc[0] == "13085"
    assert pd.isna(frame["customer_id"].iloc[1])
    # Cancellations arrive untouched: 'C' prefix and negative quantity.
    cancellation = frame[frame["invoice"] == "C579889"].iloc[0]
    assert cancellation["quantity"] == -8
    assert set(frame["source_sheet"]) == set(SOURCE_SHEETS)
    assert (frame["source_file"] == SOURCE_FILENAME).all()


def test_read_source_missing_file_names_the_url(tmp_path):
    with pytest.raises(FileNotFoundError, match=r"archive\.ics\.uci\.edu"):
        read_source(tmp_path / SOURCE_FILENAME)


def test_read_source_rejects_unexpected_columns(tmp_path):
    path = tmp_path / SOURCE_FILENAME
    wb = Workbook()
    ws = wb.active
    ws.title = SOURCE_SHEETS[0]
    ws.append(["Invoice", "Wrong"])
    wb.save(path)
    with pytest.raises(ValueError, match="unexpected columns"):
        read_source(path)


def test_ingest_bronze_end_to_end(settings):
    result = ingest_bronze(settings)

    assert result.rows_loaded == 6
    con = duckdb.connect(str(settings.duckdb_path), read_only=True)
    try:
        # Exact copy: the inter-sheet duplicate is kept (dedup is silver's job).
        dup_count = con.execute(
            "SELECT COUNT(*) FROM bronze.transactions WHERE invoice = '491200'"
        ).fetchone()[0]
        assert dup_count == 2
        types = dict(
            con.execute(
                """
                SELECT column_name, data_type FROM information_schema.columns
                WHERE table_schema = 'bronze' AND table_name = 'transactions'
                """
            ).fetchall()
        )
        assert types["invoice"] == "VARCHAR"
        assert types["customer_id"] == "VARCHAR"
        assert types["invoice_date"] == "TIMESTAMP"
        assert types["quantity"] == "BIGINT"
        assert types["loaded_at"] == "TIMESTAMP"
    finally:
        con.close()

    # Parquet export mirrors the table.
    exported = duckdb.sql(f"SELECT COUNT(*) FROM '{result.parquet_path}'").fetchone()[0]
    assert exported == 6

    dictionary = result.dictionary_path.read_text()
    assert "bronze.transactions" in dictionary
    assert "`customer_id`" in dictionary
    assert "Rows:** 6" in dictionary


def test_ingest_bronze_is_idempotent(settings):
    first = ingest_bronze(settings)
    second = ingest_bronze(settings)
    assert first.rows_loaded == second.rows_loaded == 6


def test_paths_derive_from_settings(tmp_path):
    s = Settings(_env_file=None, data_dir=Path("/tmp/cl"))
    assert s.raw_dir == Path("/tmp/cl/raw")
    assert s.duckdb_path == Path("/tmp/cl/warehouse.duckdb")
